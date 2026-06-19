import os

from dotenv import load_dotenv
load_dotenv()

import cloudinary
import cloudinary.uploader

from werkzeug.utils import secure_filename
from PIL import Image
from io import BytesIO
from flask import send_file
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.platypus import Image as ReportLabImage
from flask import Flask, render_template, request, redirect, session, send_file, jsonify
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, request, redirect, session
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    session,
    flash
)

from app.database_sqlite import (
    init_db,
    add_student,
    get_students,
    delete_student,
    update_student,
    get_users,
    add_user,
    delete_user,
    update_user_role,
    total_students,
    total_users,
    recent_students,
    get_student_by_roll,
    update_student_profile,
    mark_attendance,
    get_attendance,
    attendance_summary,
    add_marks,
    get_marks,
    marks_summary,
    add_fee,
    get_fees,
    fees_summary,
    admin_exists,
    create_organization,
    organization_exists,
    dashboard_summary,
    add_activity_log,
    get_recent_activity_logs,
    change_password,
    check_user_password,
    get_students_paginated,
    get_all_activity_logs,
    update_student_full,
    get_organization_by_id,
    update_organization_name,
    approve_user,
    reject_user,
    get_organization_by_name,
    delete_organization_data,
)

from web.auth import (
    login_user,
    is_logged_in,
    is_admin,
    logout_user
)


app = Flask(__name__)
cloudinary.config(
    cloud_name=os.environ.get("CLOUDINARY_CLOUD_NAME"),
    api_key=os.environ.get("CLOUDINARY_API_KEY"),
    api_secret=os.environ.get("CLOUDINARY_API_SECRET")
)
print("Cloud Name:", os.environ.get("CLOUDINARY_CLOUD_NAME"))
print("API Key Exists:", bool(os.environ.get("CLOUDINARY_API_KEY")))
print("API Secret Exists:", bool(os.environ.get("CLOUDINARY_API_SECRET")))
UPLOAD_FOLDER = "web/static/uploads/students"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.secret_key = "mysecretkey"
API_KEY = "student-erp-secret-key"

init_db()
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
#reset_admin_password()

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def api_key_required():
    key = request.headers.get("X-API-KEY")

    if key != API_KEY:
        return False

    return True

@app.route("/setup-admin", methods=["GET", "POST"])
def setup_admin():

    if organization_exists():
        return redirect("/login")

    if request.method == "POST":
        org_name = request.form["org_name"]
        username = request.form["username"]
        password = request.form["password"]

        org_id = create_organization(org_name)

        add_user(username, password, org_id)

        users = get_users(org_id)

        for user in users:
            if user[1] == username:
                update_user_role(user[0], "admin")
                break

        return redirect("/login")

    return render_template("setup_admin.html")

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":
        org_name = request.form["organization_name"]
        username = request.form["username"]
        password = request.form["password"]
        existing_org = get_organization_by_name(org_name)

        if existing_org:
            flash("Organization name already exists. Please choose another name.", "danger")
            return redirect("/register")
        organization_id = create_organization(org_name)

        add_user(
    username,
    password,
    organization_id,
    "admin",
    "approved"
)

        users = get_users(organization_id)

        for user in users:
            if user[1] == username:
                update_user_role(user[0], "admin")
                break

        flash("Organization registered successfully! Please login.", "success")

        return redirect("/login")

    return render_template("register.html")

@app.route("/staff-register", methods=["GET", "POST"])
def staff_register():

    if request.method == "POST":
        org_name = request.form["organization_name"]
        username = request.form["username"]
        password = request.form["password"]
        from app.database_sqlite import get_organization_by_name
        org = get_organization_by_name(org_name)

        if org is None:
            return "Organization not found"

        add_user(
            username,
            password,
            org[0],
            "staff",
            "pending"
        )

        flash("Staff account request sent. Wait for admin approval.", "warning")

        return redirect("/login")

    return render_template("role_register.html", role="Staff")


@app.route("/teacher-register", methods=["GET", "POST"])
def teacher_register():

    if request.method == "POST":
        org_name = request.form["organization_name"]
        username = request.form["username"]
        password = request.form["password"]
        from app.database_sqlite import get_organization_by_name
        org = get_organization_by_name(org_name)

        if org is None:
            return "Organization not found"

        add_user(
            username,
            password,
            org[0],
            "teacher",
            "pending"
        )

        flash("Teacher account request sent. Wait for admin approval.", "warning")

        return redirect("/login")

    return render_template("role_register.html", role="Teacher")

# 🔐 LOGIN PAGE
@app.route("/login", methods=["GET", "POST"])
def login():

    if not organization_exists():
        return redirect("/setup-admin")

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        login_result = login_user(username, password)

        if login_result == True:

            org = get_organization_by_id(
        session.get("organization_id")
            )

            if org:
                session["organization_name"] = org[1]

            return redirect("/")

        elif login_result == "pending":

            return "Your account is pending admin approval."

        else:

            return "Invalid Credentials"

    return render_template("login.html")


# 🔐 LOGOUT
@app.route("/logout")
def logout():
    logout_user()
    return redirect("/login")


# 🏠 HOME (PROTECTED)
@app.route("/")
def home():
    if not is_logged_in():
        return redirect("/login")
    
    org_id = session.get("organization_id")
    activity_logs = get_recent_activity_logs(org_id)
    dashboard = dashboard_summary(org_id)

    page = int(request.args.get("page", 1))
    per_page = 10
    offset = (page - 1) * per_page

    students = get_students_paginated(org_id, per_page, offset)
    student_count = total_students(org_id)
    total_pages = (student_count + per_page - 1) // per_page
    user_count = total_users(org_id)
    recent = recent_students(org_id)

    return render_template(
        "index.html",
        students=students,
        total=len(students),
        query="",
        student_count=student_count,
        user_count=user_count,
        recent=recent,
        dashboard=dashboard,
        activity_logs=activity_logs,
        page=page,
        total_pages=total_pages
    )

@app.route("/add", methods=["POST"])
def add():
    if not is_logged_in():
        return redirect("/login")

    name = request.form["name"]
    roll = int(request.form["roll"])

    photo_filename = None
    
    
    if "photo" in request.files:
        photo = request.files["photo"]

        if photo and photo.filename != "":

            try:
                result = cloudinary.uploader.upload(
                photo,
                folder="student_erp"
            )

                print("CLOUDINARY RESULT:", result)

                photo_filename = result["secure_url"]

            except Exception as e:
                print("CLOUDINARY ERROR:", str(e))
                raise

        print("Saved URL:", photo_filename)

    email = request.form.get("email")
    phone = request.form.get("phone")
    address = request.form.get("address")
    dob = request.form.get("dob")
    course = request.form.get("course")
    semester = request.form.get("semester")

    add_student(
        name,
        roll,
        session.get("organization_id"),
        email,
        phone,
        address,
        dob,
        course,
        semester,
        photo_filename
    )

    add_activity_log(
        session.get("user"),
        f"Added student {name} with roll {roll}",
        session.get("organization_id")
    )

    flash(
        "Student added successfully!",
        "success"
    )

    return redirect("/")


@app.route("/delete/<int:roll>")
def delete(roll):

    if not is_logged_in():
        return redirect("/login")

    #if not is_admin():
    #    return "⛔ Access Denied", 403

    delete_student(
    roll,
    session.get("organization_id")
)
    add_activity_log(
    session.get("user"),
    f"Deleted student with roll {roll}",
    session.get("organization_id")
)
    flash(
    "Student deleted successfully!",
    "danger"
)

    return redirect("/")


@app.route("/update", methods=["POST"])
def update():
    if not is_logged_in():
        return redirect("/login")

    roll = request.form.get("roll", "").strip()

    if not roll:
        return "Roll number missing", 400

    name = request.form["name"]
    email = request.form.get("email")
    phone = request.form.get("phone")
    address = request.form.get("address")
    dob = request.form.get("dob")
    course = request.form.get("course")
    semester = request.form.get("semester")

    org_id = session.get("organization_id")

    student = get_student_by_roll(int(roll), org_id)
    photo_filename = student[9] if student else None

    if "photo" in request.files:
        photo = request.files["photo"]
        print("Cloudinary upload starting...")
        

        if photo and photo.filename != "" and allowed_file(photo.filename):
            ext = photo.filename.rsplit(".", 1)[1].lower()
            photo_filename = secure_filename(f"student_{roll}.{ext}")

            filepath = os.path.join(app.config["UPLOAD_FOLDER"], photo_filename)

            image = Image.open(photo)
            image = image.resize((300, 300))
            image.save(filepath)

    update_student_full(
        int(roll),
        name,
        email,
        phone,
        address,
        dob,
        course,
        semester,
        org_id
    )

    update_student_profile(
    int(roll),
    email,
    phone,
    address,
    dob,
    course,
    semester,
    photo_filename,
    session.get("organization_id")
)

    add_activity_log(
        session.get("user"),
        f"Updated student {name} with roll {roll}",
        org_id
    )

    flash("Student updated successfully!", "warning")

    return redirect("/")

@app.route("/search", methods=["GET"])
def search():
    if not is_logged_in():
        return redirect("/login")

    query = request.args.get("q", "")

    students = get_students(session.get("organization_id"))

    # 🔍 filter logic
    filtered = [
        s for s in students
        if query.lower() in s[0].lower() or query in str(s[1])
    ]

    return render_template("index.html", students=filtered, total=len(students), query=query)

@app.route("/attendance")
def attendance_page():
    if not is_logged_in():
        return redirect("/login")

    org_id = session.get("organization_id")
    print("ATTENDANCE PAGE ORG ID:", org_id)

    students = get_students(org_id)
    attendance = get_attendance(org_id)
    summary = attendance_summary(org_id)

    return render_template(
        "attendance.html",
        students=students,
        attendance=attendance,
        summary=summary
    )

@app.route("/attendance/mark", methods=["POST"])
def attendance_mark():
    if not is_logged_in():
        return redirect("/login")

    roll = int(request.form["roll"])
    date = request.form["date"]
    status = request.form["status"]
    org_id = session.get("organization_id")
    print("ATTENDANCE SAVE:", roll, date, status, org_id)
    mark_attendance(roll, date, status, org_id)

    return redirect("/attendance")

@app.route("/marks")
def marks_page():
    if not is_logged_in():
        return redirect("/login")

    org_id = session.get("organization_id")

    students = get_students(org_id)
    marks = get_marks(org_id)
    summary = marks_summary(org_id)

    return render_template(
        "marks.html",
        students=students,
        marks=marks,
        summary=summary
    )

@app.route("/fees")
def fees_page():
    if not is_logged_in():
        return redirect("/login")

    org_id = session.get("organization_id")

    students = get_students(org_id)
    fees = get_fees(org_id)
    summary = fees_summary(org_id)

    return render_template(
        "fees.html",
        students=students,
        fees=fees,
        summary=summary
    )


@app.route("/fees/add", methods=["POST"])
def fees_add():
    if not is_logged_in():
        return redirect("/login")

    roll = int(request.form["roll"])
    total_fee = int(request.form["total_fee"])
    paid_amount = int(request.form["paid_amount"])
    payment_date = request.form["payment_date"]

    add_fee(
    roll,
    total_fee,
    paid_amount,
    payment_date,
    session.get("organization_id")
)

    return redirect("/fees")

@app.route("/marks/add", methods=["POST"])
def marks_add():
    if not is_logged_in():
        return redirect("/login")

    roll = int(request.form["roll"])
    subject = request.form["subject"]
    internal = int(request.form["internal"])
    external = int(request.form["external"])

    add_marks(
    roll,
    subject,
    internal,
    external,
    session.get("organization_id")
)

    return redirect("/marks")




@app.route("/student/<int:roll>")
def student_profile(roll):

    if not is_logged_in():
        return redirect("/login")

    student = get_student_by_roll(roll, session.get("organization_id"))

    if student is None:
        return "Student Not Found", 404

    return render_template(
        "student_profile.html",
        student=student
    )

@app.route("/student/photo/<int:roll>", methods=["POST"])
def upload_student_photo(roll):
    if not is_logged_in():
        return redirect("/login")

    student = get_student_by_roll(roll, session.get("organization_id"))

    if student is None:
        return "Student Not Found", 404

    if "photo" not in request.files:
        return redirect(f"/student/{roll}")

    photo = request.files["photo"]

    if photo.filename == "":
        return redirect(f"/student/{roll}")

    if photo and allowed_file(photo.filename):
        ext = photo.filename.rsplit(".", 1)[1].lower()
        filename = secure_filename(f"student_{roll}.{ext}")

        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)

        image = Image.open(photo)
        image = image.resize((300, 300))
        image.save(filepath)

        update_student_profile(
    roll,
    student[3],
    student[4],
    student[5],
    student[6],
    student[7],
    student[8],
    filename,
    session.get("organization_id")
)

    return redirect(f"/student/{roll}")

@app.route("/users")
def users():

    if not is_logged_in():
        return redirect("/login")

    if not is_admin():
        return "Access Denied", 403

    users = get_users(session.get("organization_id"))

    return render_template("users.html", users=users)

@app.route("/add_user", methods=["POST"])
def add_new_user():

    if not is_logged_in():
        return redirect("/login")

    if not is_admin():
        return "Access Denied", 403

    username = request.form["username"]
    password = request.form["password"]
    role = request.form["role"]
    if role == "admin":
        return "Only one admin allowed."

    add_user(
    username,
    password,
    session.get("organization_id"),
    role,
    "approved"
)

    # Role update
    users = get_users(session.get("organization_id"))
    for user in users:
        if user[1] == username:
            update_user_role(user[0], role)
            break

    return redirect("/users")

@app.route("/delete_user/<int:user_id>")
def remove_user(user_id):

    if not is_logged_in():
        return redirect("/login")

    if not is_admin():
        return "Access Denied", 403
    users = get_users(session.get("organization_id"))

    for u in users:
        if u[0] == user_id and u[1] == session.get("user"):
            return "You cannot delete your own account.", 403
    delete_user(user_id)

    return redirect("/users")

@app.route("/approve_user/<int:user_id>")
def approve_user_route(user_id):
    if not is_logged_in():
        return redirect("/login")

    if not is_admin():
        return "Access Denied", 403

    approve_user(user_id)

    return redirect("/users")


@app.route("/reject_user/<int:user_id>")
def reject_user_route(user_id):
    if not is_logged_in():
        return redirect("/login")

    if not is_admin():
        return "Access Denied", 403

    reject_user(user_id)

    return redirect("/users")

@app.route("/reports/students/pdf")
def student_report_pdf():
    if not is_logged_in():
        return redirect("/login")

    org_id = session.get("organization_id")
    students = get_students(org_id)

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()
    content = []

    content.append(Paragraph("Student Profile Report", styles["Title"]))
    content.append(Spacer(1, 12))

    table_data = [[
        "Name", "Roll", "Email", "Phone", "Course", "Semester"
    ]]

    for s in students:
        student = get_student_by_roll(s[1], org_id)

        if student is None:
            continue

        table_data.append([
            str(student[1]),
            str(student[2]),
            str(student[3] or "-"),
            str(student[4] or "-"),
            str(student[7] or "-"),
            str(student[8] or "-")
        ])

    table = Table(table_data)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
    ]))

    content.append(table)

    doc.build(content)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="students_report.pdf",
        mimetype="application/pdf"
    )

@app.route("/reports/students/excel")
def student_report_excel():
    if not is_logged_in():
        return redirect("/login")

    org_id = session.get("organization_id")
    students = get_students(org_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append(["Name", "Roll", "Email", "Phone", "Course", "Semester"])

    for s in students:
        student = get_student_by_roll(s[1], org_id)

        if student is None:
            continue

        ws.append([
            student[1],
            student[2],
            student[3] or "-",
            student[4] or "-",
            student[7] or "-",
            student[8] or "-"
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="students_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/reports/attendance/pdf")
def attendance_report_pdf():
    if not is_logged_in():
        return redirect("/login")

    attendance = get_attendance(
    session.get("organization_id")
)

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()
    content = []

    content.append(Paragraph("Attendance Report", styles["Title"]))
    content.append(Spacer(1, 12))

    table_data = [[
        "Student",
        "Roll",
        "Date",
        "Status"
    ]]

    for a in attendance:
        table_data.append([
            str(a[1] or "-"),
            str(a[2]),
            str(a[3]),
            str(a[4])
        ])

    table = Table(table_data)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
    ]))

    content.append(table)

    doc.build(content)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="attendance_report.pdf",
        mimetype="application/pdf"
    )

@app.route("/reports/attendance/excel")
def attendance_report_excel():
    if not is_logged_in():
        return redirect("/login")

    attendance = get_attendance(
    session.get("organization_id")
)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append(["Student", "Roll", "Date", "Status"])

    for a in attendance:
        ws.append([
            a[1] or "-",
            a[2],
            a[3],
            a[4]
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="attendance_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/reports/marks/pdf")
def marks_report_pdf():
    if not is_logged_in():
        return redirect("/login")

    marks = get_marks(
    session.get("organization_id")
)

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()
    content = []

    content.append(Paragraph("Marks Report", styles["Title"]))
    content.append(Spacer(1, 12))

    table_data = [[
        "Student",
        "Roll",
        "Subject",
        "Internal",
        "External",
        "Total"
    ]]

    for m in marks:
        table_data.append([
            str(m[1] or "-"),
            str(m[2]),
            str(m[3]),
            str(m[4]),
            str(m[5]),
            str(m[6])
        ])

    table = Table(table_data)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
    ]))

    content.append(table)

    doc.build(content)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="marks_report.pdf",
        mimetype="application/pdf"
    )

@app.route("/reports/marks/excel")
def marks_report_excel():
    if not is_logged_in():
        return redirect("/login")

    marks = get_marks(
    session.get("organization_id")
)

    wb = Workbook()
    ws = wb.active
    ws.title = "Marks"

    ws.append(["Student", "Roll", "Subject", "Internal", "External", "Total"])

    for m in marks:
        ws.append([
            m[1] or "-",
            m[2],
            m[3],
            m[4],
            m[5],
            m[6]
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="marks_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/reports/fees/pdf")
def fees_report_pdf():
    if not is_logged_in():
        return redirect("/login")

    fees = get_fees(
    session.get("organization_id")
)

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()
    content = []

    content.append(Paragraph("Fees Report", styles["Title"]))
    content.append(Spacer(1, 12))

    table_data = [[
        "Student",
        "Roll",
        "Total Fee",
        "Paid",
        "Pending",
        "Date"
    ]]

    for f in fees:
        table_data.append([
            str(f[1] or "-"),
            str(f[2]),
            f"₹{f[3]}",
            f"₹{f[4]}",
            f"₹{f[5]}",
            str(f[6])
        ])

    table = Table(table_data)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
    ]))

    content.append(table)
    doc.build(content)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="fees_report.pdf",
        mimetype="application/pdf"
    )


@app.route("/reports/fees/excel")
def fees_report_excel():
    if not is_logged_in():
        return redirect("/login")

    fees = get_fees(
    session.get("organization_id")
)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fees"

    ws.append(["Student", "Roll", "Total Fee", "Paid", "Pending", "Date"])

    for f in fees:
        ws.append([
            f[1] or "-",
            f[2],
            f[3],
            f[4],
            f[5],
            f[6]
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="fees_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/api/students")
def api_students():
    if not is_logged_in():
        return jsonify({"error": "Unauthorized"}), 401

    org_id = session.get("organization_id")
    students = get_students(org_id)

    data = []

    for s in students:
        student = get_student_by_roll(s[1], org_id)

        if student is None:
            continue

        data.append({
            "id": student[0],
            "name": student[1],
            "roll": student[2],
            "email": student[3],
            "phone": student[4],
            "address": student[5],
            "dob": student[6],
            "course": student[7],
            "semester": student[8],
            "photo": student[9],
            "created_at": student[10]
        })

    return jsonify(data)


@app.route("/api/attendance")
def api_attendance():
    if not is_logged_in():
        return jsonify({"error": "Unauthorized"}), 401

    org_id = session.get("organization_id")
    attendance = get_attendance(org_id)

    data = []

    for a in attendance:
        data.append({
            "id": a[0],
            "student": a[1],
            "roll": a[2],
            "date": a[3],
            "status": a[4]
        })

    return jsonify(data)

@app.route("/api/marks")
def api_marks():
    if not is_logged_in():
        return jsonify({"error": "Unauthorized"}), 401

    org_id = session.get("organization_id")
    marks = get_marks(org_id)

    data = []

    for m in marks:
        data.append({
            "id": m[0],
            "student": m[1],
            "roll": m[2],
            "subject": m[3],
            "internal": m[4],
            "external": m[5],
            "total": m[6]
        })

    return jsonify(data)

@app.route("/api/fees")
def api_fees():
    if not is_logged_in():
        return jsonify({"error": "Unauthorized"}), 401

    org_id = session.get("organization_id")
    fees = get_fees(org_id)

    data = []

    for f in fees:
        data.append({
            "id": f[0],
            "student": f[1],
            "roll": f[2],
            "total_fee": f[3],
            "paid_amount": f[4],
            "pending_amount": f[5],
            "payment_date": f[6]
        })

    return jsonify(data)

@app.route("/api/students", methods=["POST"])
def api_add_student():
    #if not is_logged_in():
    #   return jsonify({"error": "Unauthorized"}), 401

    if not api_key_required():
        return jsonify({"error": "Invalid or missing API key"}), 401
    
    data = request.get_json()

    name = data.get("name")
    roll = data.get("roll")

    if not name or not roll:
        return jsonify({"error": "Name and roll are required"}), 400

    add_student(
    name,
    int(roll),
    session.get("organization_id")
)

    return jsonify({
        "message": "Student added successfully",
        "student": {
            "name": name,
            "roll": int(roll)
        }
    }), 201

@app.route("/api/students/<int:roll>", methods=["PUT"])
def api_update_student(roll):
    data = request.get_json()
    if not api_key_required():
        return jsonify({"error": "Invalid or missing API key"}), 401
    name = data.get("name")

    if not name:
        return jsonify({"error": "Name is required"}), 400

    update_student(
    roll,
    name,
    session.get("organization_id")
)

    return jsonify({
        "message": "Student updated successfully",
        "roll": roll,
        "name": name
    })

@app.route("/api/students/<int:roll>", methods=["DELETE"])
def api_delete_student(roll):
    
    if not api_key_required():
        return jsonify({"error": "Invalid or missing API key"}), 401

    delete_student(
        roll,
        session.get("organization_id")
    )

    return jsonify({
        "message": "Student deleted successfully",
        "roll": roll
    })

@app.route("/student/profile/update/<int:roll>", methods=["POST"])
def update_profile_details(roll):
    if not is_logged_in():
        return redirect("/login")

    student = get_student_by_roll(roll, session.get("organization_id"))

    if student is None:
        return "Student Not Found", 404

    email = request.form["email"]
    phone = request.form["phone"]
    address = request.form["address"]
    dob = request.form["dob"]
    course = request.form["course"]
    semester = request.form["semester"]

    photo = student[9]

    update_student_profile(
    roll,
    email,
    phone,
    address,
    dob,
    course,
    semester,
    photo,
    session.get("organization_id")
)

    return redirect(f"/student/{roll}")

@app.route("/change-password", methods=["GET", "POST"])
def change_password_page():

    if not is_logged_in():
        return redirect("/login")

    error = None

    if request.method == "POST":

        old_password = request.form["old_password"]
        new_password = request.form["new_password"]
        confirm_password = request.form["confirm_password"]

        if not check_user_password(session.get("user"), old_password):
            error = "Old password is incorrect"

        elif new_password != confirm_password:
            error = "New password and confirm password do not match"

        elif len(new_password) < 6:
            error = "Password must be at least 6 characters"

        else:
            change_password(
                session.get("user"),
                new_password
            )

            add_activity_log(
                session.get("user"),
                "Changed account password",
                session.get("organization_id")
            )

            flash(
    "Password changed successfully!",
    "success"
)

            return redirect("/")

    return render_template("change_password.html", error=error)

@app.route("/student/id-card/<int:roll>")
def student_id_card(roll):
    if not is_logged_in():
        return redirect("/login")

    org_id = session.get("organization_id")
    student = get_student_by_roll(roll, org_id)

    if student is None:
        return "Student Not Found", 404

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()
    content = []

    content.append(Paragraph("Student ID Card", styles["Title"]))
    content.append(Spacer(1, 12))

    photo_path = None

    if student[9]:
        photo_path = os.path.join(
            app.config["UPLOAD_FOLDER"],
            student[9]
        )

    if photo_path and os.path.exists(photo_path):
        content.append(ReportLabImage(photo_path, width=100, height=100))
        content.append(Spacer(1, 12))

    data = [
        ["Name", student[1]],
        ["Roll", student[2]],
        ["Email", student[3] or "-"],
        ["Phone", student[4] or "-"],
        ["Course", student[7] or "-"],
        ["Semester", student[8] or "-"]
    ]

    table = Table(data)

    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))

    content.append(table)

    doc.build(content)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"student_{roll}_id_card.pdf",
        mimetype="application/pdf"
    )

@app.route("/students/import", methods=["POST"])
def import_students():
    if not is_logged_in():
        return redirect("/login")

    if not is_admin():
        return "Access Denied", 403

    file = request.files["file"]

    workbook = load_workbook(file)
    sheet = workbook.active

    org_id = session.get("organization_id")
    imported = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        roll = row[1]

        if name and roll:
            add_student(name, int(roll), org_id)
            imported += 1

    add_activity_log(
        session.get("user"),
        f"Imported {imported} students from Excel",
        org_id
    )

    return redirect("/")

@app.route("/students/template")
def download_student_template():

    if not is_logged_in():
        return redirect("/login")

    workbook = Workbook()
    sheet = workbook.active

    sheet.title = "Students"

    sheet["A1"] = "Name"
    sheet["B1"] = "Roll"

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="student_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/activity-logs")
def activity_logs_page():

    if not is_logged_in():
        return redirect("/login")

    logs = get_all_activity_logs(
        session.get("organization_id")
    )

    return render_template(
        "activity_logs.html",
        logs=logs
    )

@app.route("/profile")
def profile_page():
    if not is_logged_in():
        return redirect("/login")

    org = get_organization_by_id(
        session.get("organization_id")
    )

    return render_template(
        "profile.html",
        org=org
    )

@app.route("/delete-organization", methods=["POST"])
def delete_organization_route():
    if not is_logged_in():
        return redirect("/login")

    if not is_admin():
        return "Access Denied", 403

    org_id = session.get("organization_id")
    typed_name = request.form["organization_name"]

    org = get_organization_by_id(org_id)

    if org is None:
        return "Organization not found", 404

    if typed_name != org[1]:
        flash("Organization name does not match.", "danger")
        return redirect("/profile")

    delete_organization_data(org_id)

    session.clear()

    flash("Organization deleted successfully.", "danger")

    return redirect("/register")

@app.route("/organization/update", methods=["POST"])
def update_organization():
    if not is_logged_in():
        return redirect("/login")

    if not is_admin():
        return "Access Denied", 403

    new_name = request.form["organization_name"]

    update_organization_name(
        session.get("organization_id"),
        new_name
    )

    add_activity_log(
        session.get("user"),
        f"Updated organization name to {new_name}",
        session.get("organization_id")
    )

    flash("Organization name updated successfully!", "success")

    return redirect("/profile")

if __name__ == "__main__":
    app.run(debug=True)
    